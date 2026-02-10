"""
Экспорт данных клиентов в различные форматы
"""
import json
import os
import csv
import io
import base64
from datetime import datetime
import psycopg2

def handler(event: dict, context) -> dict:
    """API для экспорта клиентов в CSV и другие форматы"""
    method = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, Authorization, X-Authorization',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    token = event.get('headers', {}).get('X-Authorization', '').replace('Bearer ', '')
    if not token:
        return {
            'statusCode': 401,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Требуется авторизация'}),
            'isBase64Encoded': False
        }
    
    try:
        import jwt
        secret = os.environ.get('JWT_SECRET')
        payload = jwt.decode(token, secret, algorithms=['HS256'])
        organization_id = payload['organization_id']
    except:
        return {
            'statusCode': 401,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Неверный токен'}),
            'isBase64Encoded': False
        }
    
    if method != 'POST':
        return {
            'statusCode': 405,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Метод не поддерживается'}),
            'isBase64Encoded': False
        }
    
    try:
        body = json.loads(event.get('body', '{}'))
        action = body.get('action')
        
        if action == 'csv':
            return export_csv(organization_id, body)
        elif action == 'excel':
            return export_excel(organization_id, body)
        elif action == 'bitrix':
            return export_bitrix(organization_id, body)
        elif action == 'amocrm':
            return export_amocrm(organization_id, body)
        else:
            return {
                'statusCode': 400,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': 'Неизвестное действие'}),
                'isBase64Encoded': False
            }
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': str(e)}),
            'isBase64Encoded': False
        }

def export_csv(organization_id: int, body: dict) -> dict:
    """Экспорт клиентов в CSV формат"""
    quadrant = body.get('quadrant')
    matrix_id = body.get('matrix_id')
    
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()
    
    query = """
        SELECT 
            c.company_name,
            c.contact_person,
            c.email,
            c.phone,
            c.description,
            c.score_x,
            c.score_y,
            c.quadrant,
            m.name as matrix_name,
            c.created_at
        FROM clients c
        LEFT JOIN matrices m ON c.matrix_id = m.id
        WHERE c.organization_id = {org_id}
    """.format(org_id=organization_id)
    
    conditions = []
    if quadrant:
        conditions.append("c.quadrant = '{}'".format(quadrant))
    if matrix_id:
        conditions.append("c.matrix_id = {}".format(matrix_id))
    
    if conditions:
        query += " AND " + " AND ".join(conditions)
    
    query += " ORDER BY c.score_x DESC, c.score_y DESC"
    
    cur.execute(query)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'Компания', 'Контактное лицо', 'Email', 'Телефон', 'Описание',
        'Влияние (X)', 'Зрелость (Y)', 'Квадрант', 'Матрица', 'Дата создания'
    ])
    
    for row in rows:
        quadrant_names = {
            'focus': 'Фокус сейчас',
            'grow': 'Выращивать',
            'monitor': 'Мониторить',
            'archive': 'Архив'
        }
        quadrant_name = quadrant_names.get(row[7], row[7])
        
        writer.writerow([
            row[0], row[1], row[2], row[3], row[4],
            row[5], row[6], quadrant_name, row[8], row[9]
        ])
    
    csv_content = output.getvalue()
    csv_base64 = base64.b64encode(csv_content.encode('utf-8-sig')).decode('utf-8')
    
    filename = 'clients_export_{}.csv'.format(datetime.now().strftime('%Y%m%d_%H%M%S'))
    
    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        },
        'body': json.dumps({
            'filename': filename,
            'content': csv_base64,
            'total': len(rows)
        }),
        'isBase64Encoded': False
    }

def export_excel(organization_id: int, body: dict) -> dict:
    """Экспорт клиентов в Excel формат"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    quadrant = body.get('quadrant')
    matrix_id = body.get('matrix_id')
    
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()
    
    query = """
        SELECT 
            c.company_name,
            c.contact_person,
            c.email,
            c.phone,
            c.description,
            c.score_x,
            c.score_y,
            c.quadrant,
            m.name as matrix_name,
            c.created_at
        FROM clients c
        LEFT JOIN matrices m ON c.matrix_id = m.id
        WHERE c.organization_id = {org_id}
    """.format(org_id=organization_id)
    
    conditions = []
    if quadrant:
        conditions.append("c.quadrant = '{}'".format(quadrant))
    if matrix_id:
        conditions.append("c.matrix_id = {}".format(matrix_id))
    
    if conditions:
        query += " AND " + " AND ".join(conditions)
    
    query += " ORDER BY c.score_x DESC, c.score_y DESC"
    
    cur.execute(query)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    
    headers = [
        'Компания', 'Контактное лицо', 'Email', 'Телефон', 'Описание',
        'Влияние (X)', 'Зрелость (Y)', 'Квадрант', 'Матрица', 'Дата создания'
    ]
    
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    quadrant_names = {
        'focus': 'Фокус сейчас',
        'grow': 'Выращивать',
        'monitor': 'Мониторить',
        'archive': 'Архив'
    }
    
    quadrant_colors = {
        'focus': PatternFill(start_color="10B981", end_color="10B981", fill_type="solid"),
        'grow': PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid"),
        'monitor': PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid"),
        'archive': PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid")
    }
    
    for row_num, row_data in enumerate(rows, 2):
        quadrant = row_data[7]
        quadrant_name = quadrant_names.get(quadrant, quadrant)
        
        ws.cell(row=row_num, column=1, value=row_data[0])
        ws.cell(row=row_num, column=2, value=row_data[1])
        ws.cell(row=row_num, column=3, value=row_data[2])
        ws.cell(row=row_num, column=4, value=row_data[3])
        ws.cell(row=row_num, column=5, value=row_data[4])
        ws.cell(row=row_num, column=6, value=row_data[5])
        ws.cell(row=row_num, column=7, value=row_data[6])
        
        quadrant_cell = ws.cell(row=row_num, column=8, value=quadrant_name)
        if quadrant in quadrant_colors:
            quadrant_cell.fill = quadrant_colors[quadrant]
            quadrant_cell.font = Font(color="FFFFFF", bold=True)
            quadrant_cell.alignment = Alignment(horizontal="center")
        
        ws.cell(row=row_num, column=9, value=row_data[8])
        ws.cell(row=row_num, column=10, value=str(row_data[9]))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_content = excel_buffer.getvalue()
    excel_base64 = base64.b64encode(excel_content).decode('utf-8')
    
    filename = 'clients_export_{}.xlsx'.format(datetime.now().strftime('%Y%m%d_%H%M%S'))
    
    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        },
        'body': json.dumps({
            'filename': filename,
            'content': excel_base64,
            'total': len(rows)
        }),
        'isBase64Encoded': False
    }

def export_bitrix(organization_id: int, body: dict) -> dict:
    """Экспорт в формат Bitrix24"""
    quadrant = body.get('quadrant')
    
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()
    
    query = """
        SELECT 
            c.company_name,
            c.contact_person,
            c.email,
            c.phone,
            c.description,
            c.score_x,
            c.score_y,
            c.quadrant,
            m.name as matrix_name
        FROM clients c
        LEFT JOIN matrices m ON c.matrix_id = m.id
        WHERE c.organization_id = {org_id}
    """.format(org_id=organization_id)
    
    if quadrant:
        query += " AND c.quadrant = '{}'".format(quadrant)
    
    cur.execute(query)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    bitrix_data = []
    for row in rows:
        bitrix_data.append({
            'TITLE': row[0],
            'NAME': row[1],
            'EMAIL': [{'VALUE': row[2], 'VALUE_TYPE': 'WORK'}] if row[2] else [],
            'PHONE': [{'VALUE': row[3], 'VALUE_TYPE': 'WORK'}] if row[3] else [],
            'COMMENTS': row[4],
            'UF_CRM_SCORE_X': row[5],
            'UF_CRM_SCORE_Y': row[6],
            'UF_CRM_QUADRANT': row[7],
            'UF_CRM_MATRIX': row[8]
        })
    
    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        },
        'body': json.dumps({
            'format': 'bitrix24',
            'leads': bitrix_data,
            'total': len(bitrix_data)
        }),
        'isBase64Encoded': False
    }

def export_amocrm(organization_id: int, body: dict) -> dict:
    """Экспорт в формат amoCRM"""
    quadrant = body.get('quadrant')
    
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()
    
    query = """
        SELECT 
            c.company_name,
            c.contact_person,
            c.email,
            c.phone,
            c.description,
            c.score_x,
            c.score_y,
            c.quadrant,
            m.name as matrix_name
        FROM clients c
        LEFT JOIN matrices m ON c.matrix_id = m.id
        WHERE c.organization_id = {org_id}
    """.format(org_id=organization_id)
    
    if quadrant:
        query += " AND c.quadrant = '{}'".format(quadrant)
    
    cur.execute(query)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    amo_data = []
    for row in rows:
        amo_data.append({
            'name': row[0],
            'contacts': [{
                'name': row[1],
                'custom_fields': [
                    {'id': 'EMAIL', 'values': [{'value': row[2], 'enum': 'WORK'}]} if row[2] else None,
                    {'id': 'PHONE', 'values': [{'value': row[3], 'enum': 'WORK'}]} if row[3] else None
                ]
            }],
            'custom_fields': [
                {'id': 'DESCRIPTION', 'values': [{'value': row[4]}]},
                {'id': 'SCORE_X', 'values': [{'value': str(row[5])}]},
                {'id': 'SCORE_Y', 'values': [{'value': str(row[6])}]},
                {'id': 'QUADRANT', 'values': [{'value': row[7]}]},
                {'id': 'MATRIX', 'values': [{'value': row[8]}]}
            ]
        })
    
    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        },
        'body': json.dumps({
            'format': 'amocrm',
            'leads': amo_data,
            'total': len(amo_data)
        }),
        'isBase64Encoded': False
    }